import os
import sys
import json
import subprocess

# Ensure openpyxl is installed for Excel manipulation
try:
    import openpyxl
except ImportError:
    print("Installing 'openpyxl' to read and write Excel files...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

EXCEL_FILE = "projects.xlsx"
JSON_FILE = "projects.json"

DEFAULT_PROJECTS = [
    {
        "ID": 1,
        "Title": "Premium Relief & Wellness Listing",
        "Category": "Listing Images",
        "Client_Name": "Sombra Wellness",
        "Before_Image": "images/sombra_before.jpg",
        "After_Image": "images/sombra_after.jpg",
        "Carousel_Images": "images/sombra_after.jpg,images/sombra_infographic.jpg,images/sombra_lifestyle.jpg",
        "Metrics_Uplift": "+42% Conversion Rate",
        "Description": "Redesigned product main listing images and custom lifestyle infographics for premium pain relief gels. Solved dark packaging visibility issues by using professional studio lighting effects and text overlay callouts.",
        "Bullet_Points": "CLINICALLY PROVEN RELIEF: Targeted fast-acting formula for muscle aches and joint pains;PREMIUM NATURAL INGREDIENTS: Infused with cooling menthol and soothing aloe vera;NON-GREASY & QUICK ABSORPTION: Penetrates deep into the skin without leaving any residue;MADE IN THE USA: Formulated in a certified facility ensuring the highest safety standard"
    },
    {
        "ID": 2,
        "Title": "Outdoor String Lights EBC",
        "Category": "A+ Content",
        "Client_Name": "Banord",
        "Before_Image": "images/banord_before.jpg",
        "After_Image": "images/banord_after.jpg",
        "Carousel_Images": "images/banord_after.jpg,images/banord_module1.jpg,images/banord_module2.jpg",
        "Metrics_Uplift": "+35% Sales Volume",
        "Description": "Crafted premium A+ Content detailing the weatherproofing features and installation steps. Created dynamic night-time lifestyle scenes using AI rendering, paired with professional typography.",
        "Bullet_Points": "COMMERCIAL GRADE WEATHERPROOF: Built with heavy-duty cord and waterproof sockets;SHATTERPROOF LED BULBS: Includes energy-efficient, shatterproof plastic filament bulbs;LINKABLE DESIGN: Connect up to 40 strands to illuminate large outdoor spaces;EASY INSTALLATION: Built-in hanging loops on each socket for quick setup"
    },
    {
        "ID": 3,
        "Title": "Fluoride-Free Toothpaste A+ Content",
        "Category": "A+ Content",
        "Client_Name": "Zooba Lab",
        "Before_Image": "images/zooba_before.jpg",
        "After_Image": "images/zooba_after.jpg",
        "Carousel_Images": "images/zooba_after.jpg,images/zooba_module1.jpg,images/zooba_module2.jpg",
        "Metrics_Uplift": "+28% Conversion Rate",
        "Description": "Developed clean, organic-looking A+ modules emphasizing natural ingredients and oral health benefits. Used modern minimalist layout with custom iconography to attract eco-conscious buyers.",
        "Bullet_Points": "100% NATURAL INGREDIENTS: Organic charcoal and coconut oil for natural teeth whitening;FLUORIDE-FREE & SAFE: Gentle, non-toxic formula safe for children and daily use;FRESH PEPPERMINT BREATH: Eliminates bad breath bacteria for lasting freshness;ECO-FRIENDLY PACKAGING: Recyclable tubes protecting the environment"
    },
    {
        "ID": 4,
        "Title": "ANC Wireless Headphones Listing",
        "Category": "Listing Images",
        "Client_Name": "AudioFlex",
        "Before_Image": "images/headphones_before.jpg",
        "After_Image": "images/headphones_after.jpg",
        "Carousel_Images": "images/headphones_after.jpg,images/headphones_infographic.jpg,images/headphones_lifestyle.jpg",
        "Metrics_Uplift": "+48% Click-Through Rate",
        "Description": "Created high-impact, sleek infographics detailing Active Noise Cancellation (ANC) dB levels and battery duration. Used dark neon styling to appeal to tech enthusiasts.",
        "Bullet_Points": "ACTIVE NOISE CANCELLATION: Advanced Hybrid ANC blocking up to 35dB of external noise;60 HOURS PLAYTIME: Enjoy long-lasting music for days with fast Type-C charging;HI-RES AUDIO QUALITY: Custom-tuned 40mm dynamic drivers for deep bass and clear treble;ULTRA-COMFORT DESIGN: Breathable memory foam earcups for all-day listening comfort"
    },
    {
        "ID": 5,
        "Title": "Self Control & Focus Book",
        "Category": "Lifestyle",
        "Client_Name": "Mindful Press",
        "Before_Image": "images/book_before.jpg",
        "After_Image": "images/book_after.jpg",
        "Carousel_Images": "images/book_after.jpg,images/book_lifestyle.jpg",
        "Metrics_Uplift": "+15% Sales Increase",
        "Description": "Designed lifestyle listing images showcasing the book in a clean, quiet study environment. Added text overlays emphasizing life-changing chapters and exercises.",
        "Bullet_Points": "PRACTICAL MINDFULNESS EXERCISES: Step-by-step cognitive routines to improve daily focus;HABIT FORMATION TRACKERS: Integrated templates to build good habits and break bad ones;SCIENTIFICALLY BACKED: Written by clinical psychologists based on cognitive therapy;EASY-TO-READ FORMAT: Engaging diagrams, worksheets, and actionable summary checklists"
    },
    {
        "ID": 6,
        "Title": "Silicone Ice Cube Tray Infographics",
        "Category": "Infographics",
        "Client_Name": "IceCraft",
        "Before_Image": "images/ice_before.jpg",
        "After_Image": "images/ice_after.jpg",
        "Carousel_Images": "images/ice_after.jpg,images/ice_infographic.jpg,images/ice_lifestyle.jpg",
        "Metrics_Uplift": "+30% Conversion Rate",
        "Description": "Designed conversion-focused infographics demonstrating easy-release flexible silicone and spill-proof removable lids. Highlights durability, safety certifications, and multi-purpose freezing.",
        "Bullet_Points": "EASY RELEASE SILICONE: Soft non-stick bottom design makes removing ice cubes simple;SPILL-PROOF REMOVABLE LID: Prevents ice cubes from absorbing freezer odors and spills;FOOD-GRADE & BPA-FREE: Made of premium LFGB certified silicone for absolute safety;STACKABLE SPACE SAVER: Unique design allows stackability to save freezer space"
    }
]

def create_template_excel():
    """Create a default Excel sheet if it does not exist."""
    print(f"Creating Excel template file: {EXCEL_FILE}")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Amazon Projects"

    # Define headers
    headers = [
        "ID", "Title", "Category", "Client_Name", "Before_Image", 
        "After_Image", "Carousel_Images", "Metrics_Uplift", 
        "Description", "Bullet_Points"
    ]
    ws.append(headers)

    # Add default project rows
    for project in DEFAULT_PROJECTS:
        ws.append([
            project["ID"],
            project["Title"],
            project["Category"],
            project["Client_Name"],
            project["Before_Image"],
            project["After_Image"],
            project["Carousel_Images"],
            project["Metrics_Uplift"],
            project["Description"],
            project["Bullet_Points"]
        ])

    # Styling Excel sheet (Make header bold and apply light gray background)
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid") # Amazon Orange
    alignment = Alignment(horizontal="center", vertical="center")

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 40)

    wb.save(EXCEL_FILE)
    print("Excel template created successfully!")

def export_excel_to_json():
    """Read the Excel sheet and write to JSON."""
    if not os.path.exists(EXCEL_FILE):
        create_template_excel()

    print(f"Reading project data from {EXCEL_FILE}...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active

    projects = []
    headers = [cell.value for cell in ws[1]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: # Skip empty rows
            continue
        project = {}
        for idx, value in enumerate(row):
            if idx < len(headers):
                header = headers[idx]
                # Convert numbers if needed
                if header == "ID":
                    project[header] = int(value) if value is not None else idx
                else:
                    project[header] = str(value) if value is not None else ""
        projects.append(project)

    with open(JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(projects, f, indent=4, ensure_ascii=False)
    
    print(f"Exported {len(projects)} projects to {JSON_FILE}!")

if __name__ == "__main__":
    export_excel_to_json()
